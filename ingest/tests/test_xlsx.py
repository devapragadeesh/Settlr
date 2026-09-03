"""`.xlsx` bank-statement round-trip: generate an `.xlsx` from every real
`bank_statement.csv` on disk, using NATIVE Excel types (a `datetime.date`
object for the date cell, a `float` for the amount cell -- not text mirrors of
the CSV strings), and assert `ingest.formats.xlsx.load_bank_lines` produces
the identical `BankLine` list `ingest.formats.csv_json.load` does.

This is Phase A2 (`DECISIONS.md` Sec.81). The fixtures are SYNTHETIC --
generated from this repo's own CSVs, not a real bank's export -- and prove the
adapter is self-consistent, not that it parses an arbitrary real-world file.
Real sample files, if obtained, would be strictly better evidence; this is
what is available.
"""

from __future__ import annotations

import csv
import datetime
from pathlib import Path

import openpyxl
import pytest

from ingest.formats.csv_json import load as load_csv_json
from ingest.formats.xlsx import load_bank_lines
from resolver.loaders import _bank_column

ROOT = Path(__file__).resolve().parent.parent.parent


def _dataset_dirs() -> list[Path]:
    dirs = [ROOT / "engine" / "data", ROOT / "holdout" / "data"]
    dirs += sorted((ROOT / "scale").glob("data_*"))
    for family in ("datasets", "datasets_v2", "datasets_gst",
                    "datasets_gst_holdout", "datasets_bankside"):
        base = ROOT / "corpus" / family
        if base.exists():
            dirs += sorted(p for p in base.iterdir() if p.is_dir())
    return dirs


DATASET_DIRS = _dataset_dirs()


def _write_xlsx_from_csv(csv_path: Path, xlsx_path: Path) -> None:
    with csv_path.open(newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = reader.fieldnames or ()
        ref_col = _bank_column("reference", list(fieldnames), csv_path)
        date_col = _bank_column("value_date", list(fieldnames), csv_path)
        rows = list(reader)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Two preamble rows, as a real export would carry (account name, period)
    # -- exercises header-row detection rather than assuming row 1.
    sheet.append(["Statement export"])
    sheet.append(["Period: synthetic fixture"])
    sheet.append([ref_col, date_col, "narration", "amount"])
    for row in rows:
        value_date = datetime.date.fromisoformat(row[date_col])
        amount = float(row["amount"])
        sheet.append([row.get(ref_col) or None, value_date,
                      row.get("narration") or None, amount])
    workbook.save(xlsx_path)


@pytest.mark.parametrize("directory", DATASET_DIRS, ids=lambda d: str(d.relative_to(ROOT)))
def test_xlsx_round_trip_matches_csv(directory: Path, tmp_path: Path) -> None:
    want = load_csv_json(directory).bank

    xlsx_path = tmp_path / "bank_statement.xlsx"
    _write_xlsx_from_csv(directory / "bank_statement.csv", xlsx_path)
    got = load_bank_lines(xlsx_path)

    assert got == want


def test_a_file_with_no_recognisable_header_raises(tmp_path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for _ in range(25):
        sheet.append(["nothing", "recognisable", "here"])
    path = tmp_path / "bad.xlsx"
    workbook.save(path)
    with pytest.raises(ValueError):
        load_bank_lines(path)


def test_a_wholly_blank_trailing_row_is_skipped(tmp_path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["bank_reference", "value_date", "narration", "amount"])
    sheet.append(["REF1", datetime.date(2027, 1, 1), "note", 100.0])
    sheet.append([None, None, None, None])
    path = tmp_path / "trailing.xlsx"
    workbook.save(path)
    lines = load_bank_lines(path)
    assert len(lines) == 1
    assert lines[0].amount_paise == 10000
